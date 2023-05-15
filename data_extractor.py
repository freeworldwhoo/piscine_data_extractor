import requests
import pandas as pd
from config import token,school_link,piscine_object_name

import os

#request bearer token from /api/auth/token
token_request = requests.get(school_link+"/api/auth/token",params={"token":token})
bearer_token = token_request.json()


#get list of trgeted piscine events
headers = {"Authorization":"Bearer "+bearer_token,'Content-Type': 'application/json'}
get_piscines = {"query":'''
  {event(where:{object:{name:{_eq:"%s"}}}) {
    createdAt
    endAt
    id
    object{
      id
      name
    }
  }}
'''%(piscine_object_name)}
piscines = requests.post(school_link+"/api/graphql-engine/v1/graphql",headers=headers,json=get_piscines).json()['data']['event']


d = -1
os.system('clear')

#if no piscine event exists, exit from the script. 
if len(piscines) == 0:
  print("No piscine is ascociated with object:\t%s"%piscine_object_name)
  exit()

#loop to choose the desired event
while d < 0 or d >= len(piscines):
  for i in range(len(piscines)):
    print("%d- id: %d\n\tstart:%s\n\tend:%s\n\n"%(i,piscines[i]['id'],piscines[i]['createdAt'],piscines[i]['endAt']))
  try:
    d = int(input("what piscine index you want to get data from: "))
  except:
    pass




piscine_id = piscines[d]['id']
object_id = piscines[d]["object"]["id"]



#quesry to get all needed data
get_data = {"query":'''
  {
  levels: event_user(where: {eventId: {_eq: %d}}) {
    level
    userId
    userLogin
    user{
      attrs
    }
  }

  xp_transactions: transaction(
    where: {_and: [{type: {_eq: "xp"}}, 
      {_or: [{eventId: {_eq: %d}}, {event: {parent: {id: {_eq: %d}}}}]}]}
  ) {
    type
    amount
    user {
      login
      id
    }
  }
  
  raids_and_exams: event(where: {parent: {id: {_eq: %d}}}) {
    id
    object {
      name
      type
    }
    progresses {
      userId
      userLogin
      grade
    }
  }

  piscine_quests:object(where:{_and:[{parents:{parent:{id:{_eq:%d}}}},{type:{_eq:"quest"}}]}){
    name
    childrenRelation {
      attrs
      child{
        name
        id
      }
    }
  }

  quests_progress:progress(where:{event:{id:{_eq:%d}}}){
    grade
    object{
      name
      id
    }
    user{
      login
      id
    }
  }

  progress_on_exams:progress(
    where: {_and: [{object:{type:{_eq:"exercise"}}},{event: {object: {type: {_eq: "exam"}}}}, {event: {parent: {id: {_eq: %d}}}}]}
  ) {
    userLogin
    userId
    grade
    object {
      id
    }
    event {
      id
      object {
        name
        id
      }
    }
  }
}
'''%(piscine_id,piscine_id,piscine_id,piscine_id,object_id,piscine_id,piscine_id)
}

data = requests.post(school_link+"/api/graphql-engine/v1/graphql",headers=headers,json=get_data).json()['data']

organized_data = {}

#extarct and calculate the data to put in the dictionary
for user in data['levels']:
  organized_data[user["userId"]] = {"login":user["userLogin"],"level" : user["level"],"gender":user['user']["attrs"]["gender"]}
for transaction in data['xp_transactions']:
  userId = transaction["user"]["id"]
  if "xp" in organized_data[userId]:
    organized_data[userId]['xp'] += transaction['amount']
  else:
    organized_data[userId]['xp'] = transaction['amount']

raid_exist = False

exams_and_raids_list = []
for obj in data['raids_and_exams']:
  if obj['object']['type'] == 'exam':
    for user in obj['progresses']:
      if user['grade'] is not None:
        organized_data[user['userId']][obj['object']['name']+" [\"id: %d\"]"%obj["id"]] = round(user['grade']*100,2)
        exams_and_raids_list.append(obj['object']['name'])
  elif obj['object']['type'] == 'raid':
    for user in obj['progresses']:
      if user['grade'] is not None:
        if 'raid' in organized_data[user['userId']]:
          organized_data[user['userId']]['raid'] += int(user['grade'] >= 1)
          exams_and_raids_list.append(obj['object']['name'])
        else:
          organized_data[user['userId']]['raid'] = int(user['grade'] >= 1)
          exams_and_raids_list.append(obj['object']['name'])
        raid_exist = True
        organized_data[user['userId']][obj['object']['name']] = round(user['grade']*100,2)


quests = {}
exercises = {}
quest_exercise_name = {}
for quest in data['piscine_quests']:
  quest_exercise_name[quest['name']] = []
  quests[quest['name']] = []
  for exercise in quest['childrenRelation']:
    if not ( "category" in exercise['attrs'] and exercise['attrs']["category"] in ["optional","bonus"]):
      quests[quest['name']].append(exercise["child"]["id"])
      exercises[exercise["child"]["id"]] = quest['name']
      quest_exercise_name[quest['name']].append(exercise["child"]["name"] )

quests_user_data = {}

for id in organized_data:
  quests_user_data[id] = {}
  for quest in quests:
    quests_user_data[id][quest] = {}
    quests_user_data[id][quest]["num"] = len(quests[quest])
    quests_user_data[id][quest]["exer"] = quests[quest].copy()



for progress in data['quests_progress']:
  if progress["object"]["id"] in exercises and progress["user"]["id"] in quests_user_data and progress["object"]["id"] in quests_user_data[progress["user"]["id"]][exercises[progress["object"]["id"]]]["exer"]:
    quests_user_data[progress["user"]["id"]][exercises[progress["object"]["id"]]]["num"] -= int(progress["grade"])
    if progress["grade"] >= 1:
      quests_user_data[progress["user"]["id"]][exercises[progress["object"]["id"]]]["exer"].remove(progress["object"]["id"])

for id in organized_data:
  organized_data[id]["quests"] = sum([int(quests_user_data[id][i]["num"] == 0 ) for i in quests_user_data[id]])



exams_lvls = {}

exam_to_examlvl = {}



for prog in data['progress_on_exams']:
  exam_obj_id = prog['event']['object']['id']
  exam_event_id = prog['event']['id']
  if exam_obj_id not in exams_lvls:
    query = {"query":'''
                    {
                      object(where: {id: {_eq: %d}}) {
                        id
                        name
                        childrenRelation {
                          attrs
                          child {
                            id
                            name
                          }
                        }
                      }
                    }
    '''%(prog['event']['object']['id'])}
    exam_data = requests.post(school_link+"/api/graphql-engine/v1/graphql",headers=headers,json=query).json()['data']['object'][0]
    ex_id = exam_data['id']
    exams_lvls[ex_id] = {
        "name" : exam_data['name'],
        "max_lvl":0,
        "exercises":{}
    }
    for exercise in exam_data["childrenRelation"]:
      level = exercise['attrs']['group']
      exams_lvls[ex_id]["exercises"][exercise['child']['id']] = level
      if level > exams_lvls[ex_id]["max_lvl"]:
        exams_lvls[ex_id]["max_lvl"] = level
  if prog['grade'] >= 1:
    key = "%s  [\"id: %d\"] lvl (%d)"%(exams_lvls[exam_obj_id]['name'],exam_event_id , exams_lvls[exam_obj_id]["max_lvl"])
    exam_to_examlvl[exam_event_id] = key
    if key in organized_data[prog["userId"]]:
      organized_data[prog["userId"]][key]=max(exams_lvls[exam_obj_id]["exercises"][prog['object']['id']],organized_data[prog["userId"]][key])
    else:
      organized_data[prog["userId"]][key] = exams_lvls[exam_obj_id]["exercises"][prog['object']['id']]



#organise comlumns to extract as an excel file



data=[organized_data[id] for id in organized_data]

data_length = len(data)

columns = ["login","gender","xp","level",'quests']
if raid_exist :
  columns.append("raid")

raid_index = len(columns)

raids = []
exams = []

query = {"query":'''
                    {
                    event(where: {parent: {id: {_eq: %d}}},
                    order_by:{registration:{eventStartAt:asc}}) {
                      id
                      registration{
                        eventStartAt
                      }
                      object {
                        name
                        type
                      }
                    }
                  }
    '''%(piscine_id)}
  
events_in_order = requests.post(school_link+"/api/graphql-engine/v1/graphql",headers=headers,json=query).json()['data']['event']
for event in events_in_order:
  if event['object']['name'] in exams_and_raids_list:
    if event['object']['type'] == 'exam':
      exams.append(event['object']['name']+" [\"id: %d\"]"%event["id"] )
      exams.append(exam_to_examlvl[event["id"]])
    elif event['object']['type'] == 'raid':
      raids.append(event['object']['name'] )

columns += raids
columns += exams

table = pd.DataFrame.from_dict(data)[columns]


#extarct to excel file 
table.to_excel("%s_%d.xlsx"%(piscine_object_name,piscine_id),freeze_panes=(1,0))


#add colors to the excel file
import openpyxl
from openpyxl.styles import PatternFill,Border,Side

wb = openpyxl.load_workbook("%s_%d.xlsx"%(piscine_object_name,piscine_id)) 
ws = wb['Sheet1']

ws.delete_cols(1)

grey = PatternFill(patternType='solid', fgColor='4F5A5E')
header_color = PatternFill(patternType='solid', fgColor='80B0C8')
raid_color = PatternFill(patternType='solid', fgColor='78B7B7')
raids_color = PatternFill(patternType='solid', fgColor='20DAA5')
exam_color1 = PatternFill(patternType='solid', fgColor='C6DFE7')
exam_color2 = PatternFill(patternType='solid', fgColor='0C95C8')
login_color = PatternFill(patternType='solid', fgColor='78B7B7')
gender_color = PatternFill(patternType='solid', fgColor='AC84E8')
level_color = PatternFill(patternType='solid', fgColor='C2CCA6')
xp_color = PatternFill(patternType='solid', fgColor='E59B86')
quests_color = PatternFill(patternType='solid', fgColor='F6E5BC')

red = PatternFill(patternType='solid', fgColor='FF0000')
orange = PatternFill(patternType='solid', fgColor='F78C11')
yellow  = PatternFill(patternType='solid', fgColor='fce005')
grean = PatternFill(patternType='solid', fgColor='2FEB32')
Blue_sky = PatternFill(patternType='solid', fgColor='2FB6EB')

grade_colors = [red,orange,yellow,grean,Blue_sky,Blue_sky,Blue_sky]

thin_border = Border(left=Side(style='thin'), 
                  right=Side(style='thin'), 
                  top=Side(style='thin'), 
                  bottom=Side(style='thin'))


r = 1
for row in ws.iter_rows():
  c = 1
  for cell in row:
    if cell.internal_value is None:
      cell.fill = grey
    elif r == 1:
      cell.fill = header_color
    elif c == 1:
      cell.fill = login_color
    elif c == 2:
      cell.fill = gender_color
    elif c == 3:
      cell.fill = xp_color
    elif c == 4:
      cell.fill = level_color
    elif c == 5:
      cell.fill = quests_color
    elif c == raid_index:
      cell.fill = raid_color
    elif c > raid_index and  c <= raid_index + len(raids) :
      if cell.internal_value >= 100:
        cell.fill = Blue_sky
      else:
        cell.fill = raids_color
    elif c > raid_index + len(raids):
      if ((c - raid_index - len(raids) - 1)//2)%2 ==0:
        if (c - raid_index - len(raids) - 1) % 2 ==0:
          cell.fill = grade_colors[int(cell.internal_value)//20]
        else:
          cell.fill = exam_color1
      else:
        if (c - raid_index - len(raids) - 1) % 2 ==0:
          cell.fill = grade_colors[int(cell.internal_value)//20]
        else:
          cell.fill = exam_color2
    cell.border = thin_border  
    c+=1
  r+=1
ws.freeze_panes = 'B2'

wb.save("%s__%d.xlsx"%(piscine_object_name,piscine_id))
print("data output file name is %s"%"%s__%d.xlsx"%(piscine_object_name,piscine_id))

